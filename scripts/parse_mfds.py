import openpyxl, json, re

# Load Excel
wb = openpyxl.load_workbook('/mnt/c/Users/urtop/Downloads/20251229_음식DB 19495건.xlsx', read_only=True)
ws = wb[wb.sheetnames[0]]

# Column indices
COL_IDX = {name: i for i, name in enumerate(next(ws.iter_rows(max_row=1, values_only=True)))}
print("Key columns:", {k: COL_IDX.get(k) for k in ['식품명','데이터구분명','식품대분류명','에너지(kcal)','단백질(g)','지방(g)','탄수화물(g)']})

# Food categories of interest
TARGET_CATEGORIES = {
    '밥류': 'rice', '면 및 만두류': 'noodle', '국 및 탕류': 'soup', 
    '찌개 및 전골류': 'soup', '찜류': 'meat', '구이류': 'meat',
    '전·적 및 부침류': 'side', '볶음류': 'meat', '조림류': 'side',
    '튀김류': 'snack', '나물·숙채류': 'side', '생채·무침류': 'side',
    '김치류': 'side', '젓갈류': 'side', '장아찌·절임류': 'side',
    '반찬류': 'side', '빵 및 과자류': 'bread', '유제품류': 'side',
    '음료 및 차류': 'side', '일품요리': 'rice',
}

# Serving unit estimation based on food type
def get_servings(name, category, subcategory):
    sv = [{'unitName': 'g', 'gramsPerUnit': 1}]
    if subcategory in ('rice',):
        sv.insert(0, {'unitName': '공기', 'gramsPerUnit': 210})
        sv.insert(0, {'unitName': '그릇', 'gramsPerUnit': 400})
    elif subcategory in ('soup',):
        sv.insert(0, {'unitName': '그릇', 'gramsPerUnit': 350})
    elif subcategory in ('meat',):
        sv.insert(0, {'unitName': '인분', 'gramsPerUnit': 200})
    elif subcategory in ('noodle',):
        sv.insert(0, {'unitName': '그릇', 'gramsPerUnit': 450})
    elif subcategory in ('side',):
        sv.insert(0, {'unitName': '접시', 'gramsPerUnit': 100})
    elif subcategory in ('snack',):
        sv.insert(0, {'unitName': '개', 'gramsPerUnit': 80})
    elif subcategory in ('bread',):
        sv.insert(0, {'unitName': '조각', 'gramsPerUnit': 80})
    elif subcategory in ('fish',):
        sv.insert(0, {'unitName': '마리', 'gramsPerUnit': 150})
    return sv

# Collect foods
foods = []
seen_names = set()
stats = {'total': 0, 'filtered': 0, 'no_energy': 0}

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    stats['total'] += 1
    
    name = str(row[COL_IDX['식품명']]).strip()
    data_type = str(row[COL_IDX['데이터구분명']]).strip()
    major_cat = str(row[COL_IDX['식품대분류명']]).strip()
    
    # Only include actual foods (not raw agricultural products)
    if data_type not in ('음식', '가공식품'):
        continue
    
    # Map category
    subcategory = TARGET_CATEGORIES.get(major_cat)
    if not subcategory:
        continue
    
    stats['filtered'] += 1
    
    # Get nutrition values
    try:
        energy = float(row[COL_IDX['에너지(kcal)']] or 0)
        protein = float(row[COL_IDX['단백질(g)']] or 0)
        fat = float(row[COL_IDX['지방(g)']] or 0)
        carbs = float(row[COL_IDX['탄수화물(g)']] or 0)
    except (ValueError, TypeError):
        continue
    
    if energy <= 0:
        stats['no_energy'] += 1
        continue
    
    # Map to our category
    if '돈까스' in name or '카레' in name or '돈부리' in name or '우동' in name or '라멘' in name or '소바' in name or '초밥' in name or '사시미' in name or '타코야키' in name:
        category = 'japanese'
    elif '파스타' in name or '피자' in name or '햄버거' in name or '스테이크' in name or '샐러드' in name or '스프' in name or '리조또' in name or '오믈렛' in name or '베이컨' in name or '샌드위치' in name:
        category = 'western'
    elif '쌀국수' in name or '팟타이' in name or '똠얌' in name or '나시고렝' in name or '커리' in name or '반미' in name or '월남쌈' in name or '사테' in name or '분짜' in name:
        category = 'seasian'
    else:
        category = 'korean'
    
    # Skip duplicates
    key = name.lower()
    if key in seen_names:
        continue
    seen_names.add(key)
    
    servings = get_servings(name, category, subcategory)
    
    foods.append({
        'name': name,
        'category': category,
        'subcategory': subcategory,
        'caloriesPer100g': round(energy, 1),
        'proteinPer100g': round(protein, 1),
        'fatPer100g': round(fat, 1),
        'carbsPer100g': round(carbs, 1),
        'servings': servings,
    })

print(f"\nStats: total={stats['total']}, filtered={stats['filtered']}, no_energy={stats['no_energy']}, final={len(foods)}")

# Category breakdown
from collections import Counter
cat_count = Counter(f['category'] for f in foods)
print(f"Categories: {dict(cat_count)}")
sub_count = Counter(f['subcategory'] for f in foods)
print(f"Subcategories: {dict(sub_count)}")

# Show sample
for f in foods[:5]:
    print(f"  {f['name']} ({f['category']}/{f['subcategory']}): {f['caloriesPer100g']}kcal")

# Limit to reasonable number for seed file
# Take top foods by diversity - max 150 per category
final_foods = []
for cat in ['korean', 'japanese', 'western', 'seasian']:
    cat_foods = [f for f in foods if f['category'] == cat]
    final_foods.extend(cat_foods[:150])

# If we don't have enough in non-korean categories, fill with korean
if len(final_foods) < 300:
    remaining = 300 - len(final_foods)
    korean_foods = [f for f in foods if f['category'] == 'korean' and f not in final_foods]
    final_foods.extend(korean_foods[:remaining])

print(f"\nFinal seed count: {len(final_foods)}")

# Save JSON for seed generation
with open('/home/urtop/diet-tracker/scripts/foods_data.json', 'w', encoding='utf-8') as f:
    json.dump(final_foods, f, ensure_ascii=False, indent=2)

print("Saved to scripts/foods_data.json")
