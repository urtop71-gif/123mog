import openpyxl, sqlite3

print("📖 Loading Excel...")
wb = openpyxl.load_workbook('/mnt/c/Users/urtop/Downloads/20251229_음식DB 19495건.xlsx', read_only=True)
ws = wb[wb.sheetnames[0]]

# Build food_code -> detailed nutrients map from Excel
# Column indices: 
# 0:식품코드, 1:식품명, 17:에너지, 19:단백질, 20:지방, 22:탄수화물
# 23:당류, 25:식이섬유, 29:나트륨, 40:포화지방산, 41:트랜스지방산
COLS = {'name': 1, 'energy': 17, 'protein': 19, 'fat': 20, 'carbs': 22, 
        'sugar': 23, 'fiber': 25, 'sodium': 29, 'sat_fat': 40, 'trans_fat': 41}

food_nutrition = {}
print("📊 Building nutrition map from 19,495 foods...")
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None: continue
    name = str(row[1] or '').strip()
    try:
        food_nutrition[name.lower()] = {
            'energy': float(row[COLS['energy']] or 0),
            'protein': float(row[COLS['protein']] or 0),
            'fat': float(row[COLS['fat']] or 0),
            'carbs': float(row[COLS['carbs']] or 0),
            'sugar': float(row[COLS['sugar']] or 0),
            'fiber': float(row[COLS['fiber']] or 0),
            'sodium': float(row[COLS['sodium']] or 0),
            'sat_fat': float(row[COLS['sat_fat']] or 0),
            'trans_fat': float(row[COLS['trans_fat']] or 0),
        }
    except:
        continue

print(f"✅ {len(food_nutrition)} foods with detailed nutrients")

# Tagging functions
def tag_ldl(name, pro, fat, carb, sugar, fiber, sodium, sat_fat, trans_fat):
    """Precise LDL tagging using 포화지방, 트랜스지방, 식이섬유, 나트륨"""
    n = name.lower()
    
    # If no sat_fat data, estimate based on food type
    if sat_fat == 0 and fat > 1:
        n2 = name.lower()
        is_meat = any(w in n2 for w in ['고기', '갈비', '삼겹', '목살', '족발', '곱창', '삼계', '닭', '오리', '소시지', '베이컨', '스테이크', '버거', '치킨', '소불', '돼지', '육포', '보쌈', '수육', '등심', '안심', '갈매기', '항정'])
        is_dairy = any(w in n2 for w in ['치즈', '버터', '크림', '우유', '요거트', '생크림', '마요네즈', '라떼', '카푸치노'])
        is_fish = any(w in n2 for w in ['고등어', '꽁치', '참치', '연어', '갈치', '조기', '장어', '삼치', '전갱이', '방어', '멸치', '새우', '오징어', '문어', '조개', '굴', '홍합'])
        is_fried = any(w in n2 for w in ['튀김', '프라이', '돈까스', '커틀렛', '가라아게'])
        is_bread = any(w in n2 for w in ['빵', '케이크', '도넛', '패스츄리', '크로와상', '파이', '쿠키', '과자', '비스킷'])
        
        if is_dairy:
            sat_fat = fat * 0.6
        elif is_meat:
            sat_fat = fat * 0.45
        elif is_fried:
            sat_fat = fat * 0.35
        elif is_bread:
            sat_fat = fat * 0.4
        elif is_fish:
            sat_fat = fat * 0.2
        else:
            sat_fat = fat * 0.25
    
    # LDL BAD: high saturated fat, trans fat, high sodium + low fiber processed foods
    if sat_fat > 0 and sat_fat > 3.5:
        return 'ldl_bad'
    if trans_fat > 0 and trans_fat > 0.3:
        return 'ldl_bad'
    if sat_fat > 2 and sodium > 400 and fiber < 2:
        return 'ldl_bad'
    if sat_fat > 2.5 and fiber < 1:
        return 'ldl_bad'
    
    # Explicit junk
    if any(w in n for w in ['튀김', '프라이', '버터', '베이컨', '소시지', '마가린', '삼겹살', '족발', '곱창', '막창']):
        if sat_fat > 4:
            return 'ldl_bad'
    
    # LDL GOOD: omega-3 fish, high fiber plant foods, unsaturated fat dominant
    if any(w in n for w in ['고등어', '꽁치', '연어', '참치', '삼치', '정어리', '전갱이', '방어']):
        return 'ldl_good'
    if fiber > 0 and fiber > 3 and sat_fat < 2:
        return 'ldl_good'
    if any(w in n for w in ['아몬드', '호두', '땅콩', '캐슈', '피스타', '잣', '아보카도', '올리브', '들기름', '참기름']):
        return 'ldl_good'
    if fiber > 0 and fiber > 2 and sat_fat < 1.5 and sodium < 300:
        return 'ldl_good'
    # Very lean protein
    if fat < 3 and pro > 10 and sat_fat < 1:
        return 'ldl_good'
    
    return 'ldl_neutral'

def tag_sugar(name, pro, fat, carb, sugar, fiber, sodium, sat_fat, trans_fat):
    """Precise blood sugar tagging using 당류, 식이섬유, 탄수화물"""
    n = name.lower()
    
    # If no sugar data but high carbs with low fiber/protein, estimate
    if sugar == 0 and carb > 30 and fiber < 1:
        sugar = carb * 0.3  # estimate ~30% of carbs are sugars for refined foods
    
    # SUGAR BAD: high sugar, high refined carbs, low fiber
    if sugar > 0 and sugar > 15:
        return 'sugar_bad'
    if carb > 40 and fiber < 2 and sugar > 5:
        return 'sugar_bad'
    if carb > 50 and fiber < 1:
        return 'sugar_bad'
    if any(w in n for w in ['설탕', '시럽', '잼', '꿀', '사탕', '초콜릿', '케이크', '쿠키', '과자', '콜라', '사이다']):
        return 'sugar_bad'
    if sugar > 10 and fiber < 1:
        return 'sugar_bad'
    
    # SUGAR GOOD: low carb/sugar, high fiber, high protein
    if fiber > 0 and fiber > 3 and sugar < 5:
        return 'sugar_good'
    if carb < 10 and sugar < 3:
        return 'sugar_good'
    if pro > 15 and carb < 15:
        return 'sugar_good'
    if any(w in n for w in ['나물', '채소', '샐러드', '김치', '두부', '계란', '닭가슴살', '생선']):
        return 'sugar_good'
    if fiber > 2 and sugar < 3 and carb < 25:
        return 'sugar_good'
    
    return 'sugar_neutral'

# Update DB
conn = sqlite3.connect('dev.db')
cur = conn.cursor()
cur.execute('SELECT id, name, proteinPer100g, fatPer100g, carbsPer100g FROM Food')
foods = cur.fetchall()

stats = {'ldl': {}, 'sugar': {}, 'matched': 0, 'unmatched': 0}
for fid, name, pro, fat, carbs in foods:
    nut = food_nutrition.get(name.lower(), {})
    if nut:
        stats['matched'] += 1
        sugar = nut.get('sugar', 0)
        fiber = nut.get('fiber', 0)
        sodium = nut.get('sodium', 0)
        sat_fat = nut.get('sat_fat', 0)
        trans_fat = nut.get('trans_fat', 0)
    else:
        stats['unmatched'] += 1
        sugar = fiber = sodium = sat_fat = trans_fat = 0
    
    ldl = tag_ldl(name, pro or 0, fat or 0, carbs or 0, sugar, fiber, sodium, sat_fat, trans_fat)
    sug = tag_sugar(name, pro or 0, fat or 0, carbs or 0, sugar, fiber, sodium, sat_fat, trans_fat)
    
    stats['ldl'][ldl] = stats['ldl'].get(ldl, 0) + 1
    stats['sugar'][sug] = stats['sugar'].get(sug, 0) + 1
    
    cur.execute('UPDATE Food SET healthTags = ? WHERE id = ?', (f'{ldl},{sug}', fid))

conn.commit()
conn.close()

print(f"\n🏷️ 정밀 태깅 완료! ({stats['matched']} foods with 식약처 detail, {stats['unmatched']} with basic data)")
print(f"\nLDL:")
for k, v in sorted(stats['ldl'].items()):
    emoji = '👍' if 'good' in k else '👎' if 'bad' in k else '💛'
    print(f'  {emoji} {k}: {v}종')
print(f"\n혈당:")
for k, v in sorted(stats['sugar'].items()):
    emoji = '👍' if 'good' in k else '👎' if 'bad' in k else '💛'
    print(f'  {emoji} {k}: {v}종')
